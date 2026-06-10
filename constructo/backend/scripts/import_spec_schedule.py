"""Import a Material Specification Schedule .xlsx into the Spec engine.

SAFE BY DEFAULT — a bare run only parses and prints a plan (no DB writes). Add
``--run`` to write. ``--purge`` removes everything a prior import created for the
site. Idempotent (deterministic uuid5), so re-running never duplicates.

Examples
--------
# Dry-run — parse + print what would be created:
    uv run python -m scripts.import_spec_schedule \
        --file "/path/Interior Material Specification Template-Anil Sir.xlsx" \
        --sheet "Material Specifications" --company <COMPANY_UUID> --site <SITE_UUID>

# Actually import:
    uv run python -m scripts.import_spec_schedule --file "..." --sheet "Material Specifications" \
        --company <COMPANY_UUID> --site <SITE_UUID> --run

# Wipe what this import created for the site:
    uv run python -m scripts.import_spec_schedule \
        --company <COMPANY_UUID> --site <SITE_UUID> --purge
"""
from __future__ import annotations

import argparse
import asyncio
from uuid import UUID

from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models import Component, Space, Spec
from app.specs.import_parser import parse_spec_sheet
from app.specs.importer import import_rows


def _read_sheet(path: str, sheet: str) -> list[list]:
    import openpyxl  # imported here so parser/importer tests don't need openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


async def _purge(company_id: UUID, site_id: UUID) -> None:
    async with SessionLocal() as session:
        await session.execute(delete(Spec).where(Spec.site_id == site_id))
        result = await session.execute(select(Space.id).where(Space.site_id == site_id))
        space_ids = result.scalars().all()
        if space_ids:
            await session.execute(delete(Component).where(Component.space_id.in_(space_ids)))
        await session.execute(delete(Space).where(Space.site_id == site_id))
        # Materials are company-wide catalog; leave them (shared). Comment in if you want them gone:
        # await session.execute(delete(Material).where(Material.company_id == company_id))
        await session.commit()
    print(f"Purged specs/components/spaces for site {site_id}.")


async def _main() -> None:
    ap = argparse.ArgumentParser(description="Import a Material Specification Schedule .xlsx")
    ap.add_argument("--file")
    ap.add_argument("--sheet", default="Material Specifications")
    ap.add_argument("--company", required=True, type=UUID)
    ap.add_argument("--site", required=True, type=UUID)
    ap.add_argument("--run", action="store_true", help="write to DB (default: dry-run)")
    ap.add_argument("--purge", action="store_true", help="remove this import for the site")
    args = ap.parse_args()

    if args.purge:
        await _purge(args.company, args.site)
        return
    if not args.file:
        ap.error("--file is required unless --purge")

    parsed = parse_spec_sheet(_read_sheet(args.file, args.sheet))
    print(f"Parsed {len(parsed)} spec rows from '{args.sheet}'.")
    by_room: dict[str, int] = {}
    for p in parsed:
        by_room[p.room] = by_room.get(p.room, 0) + 1
    for room, n in by_room.items():
        print(f"  {room}: {n} lines")

    if not args.run:
        print("\nDRY RUN — no DB writes. Re-run with --run to import.")
        return

    async with SessionLocal() as session:
        stats = await import_rows(session, args.company, args.site, parsed)
        await session.commit()
    print(f"\nImported: {stats.spaces} spaces, {stats.components} components, "
          f"{stats.materials} materials, {stats.specs} specs.")


if __name__ == "__main__":
    asyncio.run(_main())
